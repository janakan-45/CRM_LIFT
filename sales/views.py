from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from .models import Customer, Route, Branch, ProvinceState, Quotation,Invoice,RecurringInvoice,RecurringInvoiceItem,PaymentReceived,InvoiceItem
from .serializers import CustomerSerializer, RouteSerializer, BranchSerializer, ProvinceStateSerializer, QuotationSerializer,InvoiceSerializer,RecurringInvoiceItemSerializer,RecurringInvoiceSerializer,PaymentReceivedSerializer
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from rest_framework.permissions import BasePermission


###########################################Customer  views######################################### 

class IsOwner(BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.groups.filter(name='owner').exists()


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsOwner])
def check_owner_status(request):
    user = request.user
    return Response({
        'message': f'User {user.username} is an owner',
        'is_owner': True
    }, status=status.HTTP_200_OK)

# Dynamic dropdown views
@api_view(['POST'])
@permission_classes([IsAuthenticated,IsOwner])
def add_route(request):
    serializer = RouteSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated,IsOwner])
def edit_route(request, pk):
    try:
        route = Route.objects.get(pk=pk)
    except Route.DoesNotExist:
        return Response({"error": "Route not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = RouteSerializer(route, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Route updated successfully!",
            "route_id": route.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated,IsOwner])
def delete_route(request, pk):
    try:
        route = Route.objects.get(pk=pk)
        route.delete()
        return Response({"message": "Route deleted successfully!"}, status=status.HTTP_200_OK)
    except Route.DoesNotExist:
        return Response({"error": "Route not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([IsAuthenticated,IsOwner])
def add_branch(request):
    serializer = BranchSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated,IsOwner])
def edit_branch(request, pk):
    try:
        branch = Branch.objects.get(pk=pk)
    except Branch.DoesNotExist:
        return Response({"error": "Branch not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = BranchSerializer(branch, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Branch updated successfully!",
            "branch_id": branch.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated,IsOwner])
def delete_branch(request, pk):
    try:
        branch = Branch.objects.get(pk=pk)
        branch.delete()
        return Response({"message": "Branch deleted successfully!"}, status=status.HTTP_200_OK)
    except Branch.DoesNotExist:
        return Response({"error": "Branch not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated,IsOwner])
def add_province_state(request):
    serializer = ProvinceStateSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated,IsOwner])
def edit_province_state(request, pk):
    try:
        province_state = ProvinceState.objects.get(pk=pk)
    except ProvinceState.DoesNotExist:
        return Response({"error": "ProvinceState not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = ProvinceStateSerializer(province_state, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "ProvinceState updated successfully!",
            "province_state_id": province_state.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated,IsOwner])
def delete_province_state(request, pk):
    try:
        province_state = ProvinceState.objects.get(pk=pk)
        province_state.delete()
        return Response({"message": "ProvinceState deleted successfully!"}, status=status.HTTP_200_OK)
    except ProvinceState.DoesNotExist:
        return Response({"error": "ProvinceState not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_routes(request):
    routes = Route.objects.all()
    serializer = RouteSerializer(routes, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_branches(request):
    branches = Branch.objects.all()
    serializer = BranchSerializer(branches, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_province_states(request):
    province_states = ProvinceState.objects.all()
    serializer = ProvinceStateSerializer(province_states, many=True)
    return Response(serializer.data)

# Customer CRUD views
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_customer(request):
    serializer = CustomerSerializer(data=request.data)
    if serializer.is_valid():
        customer = serializer.save()
        return Response({
            "message": "Customer added successfully!",
            "reference_id": customer.reference_id,
            "site_id": customer.site_id
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_customer(request, pk):
    try:
        customer = Customer.objects.get(pk=pk)
    except Customer.DoesNotExist:
        return Response({"error": "Customer not found"}, status=status.HTTP_404_NOT_FOUND)

    serializer = CustomerSerializer(customer, data=request.data, partial=True)
    if serializer.is_valid():
        customer = serializer.save()
        return Response({
            "message": "Customer updated successfully!",
            "reference_id": customer.reference_id,
            "site_id": customer.site_id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_customer(request, pk):
    try:
        customer = Customer.objects.get(pk=pk)
        customer.delete()
        return Response({"message": "Customer deleted successfully!"}, status=status.HTTP_200_OK)
    except Customer.DoesNotExist:
        return Response({"error": "Customer not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def customer_list(request):
    customers = Customer.objects.all()
    serializer = CustomerSerializer(customers, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_customers_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = [
        'Reference ID', 'Site ID', 'Site Name', 'Contact Person Name', 'Email', 'Site Address',
        'Route', 'Branch', 'Province/State', 'Sector', 'Uploads Files'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    customers = Customer.objects.all()

    for row_num, customer in enumerate(customers, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = customer.reference_id
        ws[f"{get_column_letter(2)}{row_num}"] = customer.site_id
        ws[f"{get_column_letter(3)}{row_num}"] = customer.site_name
        ws[f"{get_column_letter(4)}{row_num}"] = customer.contact_person_name
        ws[f"{get_column_letter(5)}{row_num}"] = customer.email
        ws[f"{get_column_letter(6)}{row_num}"] = customer.site_address
        ws[f"{get_column_letter(7)}{row_num}"] = customer.routes.value if customer.routes else ''
        ws[f"{get_column_letter(8)}{row_num}"] = customer.branch.value if customer.branch else ''
        ws[f"{get_column_letter(9)}{row_num}"] = customer.province_state.value if customer.province_state else ''
        ws[f"{get_column_letter(10)}{row_num}"] = customer.sector
        ws[f"{get_column_letter(11)}{row_num}"] = str(customer.uploads_files) if customer.uploads_files else ''

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customers_export.xlsx'
    response.write(output.read())

    return response

import csv
import io
from django.utils import timezone

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_customers_csv(request):
    if 'file' not in request.FILES:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

    csv_file = request.FILES['file']
    if not csv_file.name.endswith('.csv'):
        return Response({"error": "File is not a CSV"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        decoded_file = csv_file.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        reader = csv.reader(io_string, delimiter=',')
        next(reader)  # Skip header row
        for row in reader:
            if not row:  # Skip blank rows
                continue
            if any(',' in field for field in row):  # Check for commas in data
                return Response({"error": "CSV contains commas in data"}, status=status.HTTP_400_BAD_REQUEST)

            # Map CSV columns to Customer model fields
            # Assuming CSV order: site_id, job_no, site_name, site_address, email, phone, office_address,
            # contact_person_name, designation, pin_code, country, province_state_value, city, sector,
            # routes_value, branch_value, handover_date, billing_name, pan_number, gst_number, uploads_files
            customer_data = {
                'site_id': row[0] if row[0] else '',
                'job_no': row[1] if row[1] else '',
                'site_name': row[2] if row[2] else '',
                'site_address': row[3] if row[3] else '',
                'email': row[4] if row[4] else '',
                'phone': row[5] if row[5] else '',
                'office_address': row[6] if row[6] else '',
                'contact_person_name': row[7] if row[7] else '',
                'designation': row[8] if row[8] else '',
                'pin_code': row[9] if row[9] else '',
                'country': row[10] if row[10] else '',
                'province_state': ProvinceState.objects.get_or_create(value=row[11])[0] if row[11] else None,
                'city': row[12] if row[12] else '',
                'sector': row[13] if row[13] in ['government', 'private'] else 'private',
                'routes': Route.objects.get_or_create(value=row[14])[0] if row[14] else None,
                'branch': Branch.objects.get_or_create(value=row[15])[0] if row[15] else None,
                'handover_date': timezone.datetime.strptime(row[16], '%Y-%m-%d').date() if row[16] else None,
                'billing_name': row[17] if row[17] else '',
                'pan_number': row[18] if row[18] else '',
                'gst_number': row[19] if row[19] else '',
                'active_mobile': int(row[20]) if row[20] else 0,
                'expired_mobile': int(row[21]) if row[21] else 0,
                'contracts': int(row[22]) if row[22] else 0,
                'no_of_lifts': int(row[23]) if row[23] else 0,
                'completed_services': int(row[24]) if row[24] else 0,
                'due_services': int(row[25]) if row[25] else 0,
                'overdue_services': int(row[26]) if row[26] else 0,
                'tickets': int(row[27]) if row[27] else 0,
                'uploads_files': row[28] if row[28] else None,
            }
            serializer = CustomerSerializer(data=customer_data)
            if serializer.is_valid():
                serializer.save()
            else:
                return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"message": "Customers imported successfully!"}, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
###########################################quotation###################################3



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_quotation(request):
    serializer = QuotationSerializer(data=request.data)
    if serializer.is_valid():
        quotation = serializer.save()
        return Response({
            "message": "Quotation added successfully!",
            "quotation_id": quotation.id
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_quotation(request, pk):
    try:
        quotation = Quotation.objects.get(pk=pk)
    except Quotation.DoesNotExist:
        return Response({"error": "Quotation not found"}, status=status.HTTP_404_NOT_FOUND)

    serializer = QuotationSerializer(quotation, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Quotation updated successfully!",
            "quotation_id": quotation.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_quotation(request, pk):
    try:
        quotation = Quotation.objects.get(pk=pk)
        quotation.delete()
        return Response({"message": "Quotation deleted successfully!"}, status=status.HTTP_200_OK)
    except Quotation.DoesNotExist:
        return Response({"error": "Quotation not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def quotation_list(request):
    quotations = Quotation.objects.all()
    serializer = QuotationSerializer(quotations, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_quotations_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotations"

    headers = [
        'Reference ID', 'Customer', 'AMC Type', 'Sales/Service Executive', 'Type',
        'Year of Make', 'Date', 'Remark', 'Other Remark'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    quotations = Quotation.objects.all()

    for row_num, quotation in enumerate(quotations, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = quotation.reference_id
        ws[f"{get_column_letter(2)}{row_num}"] = quotation.customer.site_name if quotation.customer else ''
        ws[f"{get_column_letter(3)}{row_num}"] = quotation.amc_type.name if quotation.amc_type else ''
        ws[f"{get_column_letter(4)}{row_num}"] = quotation.sales_service_executive.name if quotation.sales_service_executive else ''
        ws[f"{get_column_letter(5)}{row_num}"] = quotation.type
        ws[f"{get_column_letter(6)}{row_num}"] = quotation.year_of_make
        ws[f"{get_column_letter(7)}{row_num}"] = quotation.date.strftime('%Y-%m-%d')
        ws[f"{get_column_letter(8)}{row_num}"] = quotation.remark
        ws[f"{get_column_letter(9)}{row_num}"] = quotation.other_remark

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=quotations_export.xlsx'
    response.write(output.read())

    return response




############################invoice views#########################################



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_invoice(request):
    serializer = InvoiceSerializer(data=request.data)
    if serializer.is_valid():
        invoice = serializer.save()
        return Response({
            "message": "Invoice added successfully!",
            "reference_id": invoice.reference_id
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_invoice(request, pk):
    try:
        invoice = Invoice.objects.get(pk=pk)
    except Invoice.DoesNotExist:
        return Response({"error": "Invoice not found"}, status=status.HTTP_404_NOT_FOUND)

    serializer = InvoiceSerializer(invoice, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Invoice updated successfully!",
            "reference_id": invoice.reference_id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_invoice(request, pk):
    try:
        invoice = Invoice.objects.get(pk=pk)
        invoice.delete()
        return Response({"message": "Invoice deleted successfully!"}, status=status.HTTP_200_OK)
    except Invoice.DoesNotExist:
        return Response({"error": "Invoice not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def invoice_list(request):
    invoices = Invoice.objects.all()
    serializer = InvoiceSerializer(invoices, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_invoices_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        'Reference ID', 'Customer', 'AMC Type', 'Start Date', 'Due Date',
        'Discount', 'Payment Term', 'Status'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    invoices = Invoice.objects.all()

    for row_num, invoice in enumerate(invoices, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = invoice.reference_id
        ws[f"{get_column_letter(2)}{row_num}"] = invoice.customer_name if invoice.customer_name else ''
        ws[f"{get_column_letter(3)}{row_num}"] = invoice.amc_type_name if invoice.amc_type_name else ''
        ws[f"{get_column_letter(4)}{row_num}"] = invoice.start_date.strftime('%Y-%m-%d')
        ws[f"{get_column_letter(5)}{row_num}"] = invoice.due_date.strftime('%Y-%m-%d')
        ws[f"{get_column_letter(6)}{row_num}"] = str(invoice.discount)
        ws[f"{get_column_letter(7)}{row_num}"] = invoice.payment_term
        ws[f"{get_column_letter(8)}{row_num}"] = invoice.status

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=invoices_export.xlsx'
    response.write(output.read())

    return response


# views.py (add this at the end or appropriate place)

import logging
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from io import BytesIO
from django.http import HttpResponse

logger = logging.getLogger(__name__)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def print_invoice(request, pk):
    try:
        invoice = Invoice.objects.get(pk=pk)
    except Invoice.DoesNotExist:
        return Response({"error": "Invoice not found"}, status=status.HTTP_404_NOT_FOUND)

    try:
        # Prepare data
        context = {
            'company_name': 'Atom Lifts India Pvt Ltd',
            'address': 'No.87B, Pillayar Koll Street, Mannurpet, Ambattur Indus Estate, Chennai 50, CHENNAI',
            'phone': '9600087456',
            'email': 'admin@atomlifts.com',
            'invoice_no': invoice.reference_id,
            'invoice_date': invoice.start_date.strftime('%d/%m/%Y') if invoice.start_date else 'N/A',
            'due_date': invoice.due_date.strftime('%d/%m/%Y') if invoice.due_date else 'N/A',
            'customer_name': invoice.customer.site_name if invoice.customer else 'N/A',
            'customer_address': invoice.customer.site_address if invoice.customer else 'N/A',
            'customer_email': invoice.customer.email if invoice.customer else 'N/A',
            'customer_phone': invoice.customer.phone if invoice.customer else 'N/A',
            'discount': str(invoice.discount),
            'payment_term': invoice.payment_term.upper(),
            'status': invoice.status.upper(),
            'items': [],
            'subtotal': 0,
            'tax_total': 0,
            'grand_total': 0,
        }

        # Process items
        for item in invoice.items.all():
            item_total = item.rate * item.qty
            item_tax_amount = item_total * (item.tax / 100)
            item_grand = item_total + item_tax_amount
            context['items'].append({
                'item_name': item.item.name if item.item else 'N/A',
                'rate': str(item.rate),
                'qty': str(item.qty),
                'tax': str(item.tax),
                'total': str(item_total),
                'tax_amount': str(item_tax_amount),
                'grand': str(item_grand),
            })
            context['subtotal'] += item_total
            context['tax_total'] += item_tax_amount
            context['grand_total'] += item_grand

        # Apply discount to grand total
        discount_amount = context['grand_total'] * (float(invoice.discount) / 100)
        context['grand_total'] -= discount_amount
        context['discount_amount'] = str(discount_amount)

        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
        styles = getSampleStyleSheet()
        story = []

        # Header
        header_style = ParagraphStyle(
            name='HeaderStyle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1  # Center
        )
        story.append(Paragraph(context['company_name'], header_style))
        story.append(Paragraph(context['address'], styles['Normal']))
        story.append(Paragraph(f"Phone: {context['phone']} | Email: {context['email']}", styles['Normal']))
        story.append(Spacer(1, 12))

        # Invoice Details Table
        story.append(Paragraph("Invoice Details", styles['Heading2']))
        data = [
            ['Invoice No:', context['invoice_no']],
            ['Invoice Date:', context['invoice_date']],
            ['Due Date:', context['due_date']],
            ['Payment Term:', context['payment_term']],
            ['Status:', context['status']],
        ]
        table = Table(data, colWidths=[150, 350])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

        # Customer Details
        story.append(Paragraph("Customer Details", styles['Heading2']))
        data = [
            ['Customer Name:', context['customer_name']],
            ['Address:', context['customer_address']],
            ['Email:', context['customer_email']],
            ['Phone:', context['customer_phone']],
        ]
        table = Table(data, colWidths=[150, 350])
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

        # Items Table
        story.append(Paragraph("Items", styles['Heading2']))
        if context['items']:
            data = [['Item Name', 'Rate', 'Qty', 'Tax (%)', 'Subtotal', 'Tax Amount', 'Total']]
            for item in context['items']:
                data.append([
                    item['item_name'],
                    item['rate'],
                    item['qty'],
                    item['tax'],
                    item['total'],
                    item['tax_amount'],
                    item['grand']
                ])
            table = Table(data, colWidths=[150, 60, 40, 60, 60, 60, 60])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No items found.", styles['Normal']))
        story.append(Spacer(1, 12))

        # Totals
        story.append(Paragraph("Totals", styles['Heading2']))
        data = [
            ['Subtotal:', str(context['subtotal'])],
            ['Tax Total:', str(context['tax_total'])],
            ['Discount (%):', context['discount']],
            ['Discount Amount:', context['discount_amount']],
            ['Grand Total:', str(context['grand_total'])],
        ]
        table = Table(data, colWidths=[150, 350])
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(table)

        # Build PDF
        doc.build(story)

        # Return PDF response
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=invoice_{context["invoice_no"]}.pdf'
        return response

    except Exception as e:
        logger.error(f"Unexpected error in print_invoice: {str(e)}")
        return Response({"error": f"Unexpected error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


###########################################recurring invoice views#########################################

# Add the following to views.py at the end, after the invoice views

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def add_recurring_invoice(request):
#     serializer = RecurringInvoiceSerializer(data=request.data)
#     if serializer.is_valid():
#         recurring_invoice = serializer.save()
#         return Response({
#             "message": "Recurring Invoice added successfully!",
#             "reference_id": recurring_invoice.reference_id
#         }, status=status.HTTP_201_CREATED)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# def edit_recurring_invoice(request, pk):
#     try:
#         recurring_invoice = RecurringInvoice.objects.get(pk=pk)
#     except RecurringInvoice.DoesNotExist:
#         return Response({"error": "Recurring Invoice not found"}, status=status.HTTP_404_NOT_FOUND)
#     serializer = RecurringInvoiceSerializer(recurring_invoice, data=request.data, partial=True)
#     if serializer.is_valid():
#         serializer.save()
#         return Response({
#             "message": "Recurring Invoice updated successfully!",
#             "reference_id": recurring_invoice.reference_id
#         }, status=status.HTTP_200_OK)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def delete_recurring_invoice(request, pk):
#     try:
#         recurring_invoice = RecurringInvoice.objects.get(pk=pk)
#         recurring_invoice.delete()
#         return Response({"message": "Recurring Invoice deleted successfully!"}, status=status.HTTP_200_OK)
#     except RecurringInvoice.DoesNotExist:
#         return Response({"error": "Recurring Invoice not found"}, status=status.HTTP_404_NOT_FOUND)

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def recurring_invoice_list(request):
#     recurring_invoices = RecurringInvoice.objects.all()
#     serializer = RecurringInvoiceSerializer(recurring_invoices, many=True)
#     return Response(serializer.data)

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def export_recurring_invoices_to_excel(request):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Recurring Invoices"

#     headers = [
#         'Reference ID', 'Customer', 'Profile Name', 'Order Number', 'Repeat Every',
#         'Start Date', 'End Date', 'Sales Person', 'Billing Address', 'GST Treatment'
#     ]

#     for col_num, header in enumerate(headers, 1):
#         ws[f"{get_column_letter(col_num)}1"] = header

#     recurring_invoices = RecurringInvoice.objects.all()

#     for row_num, invoice in enumerate(recurring_invoices, 2):
#         ws[f"{get_column_letter(1)}{row_num}"] = invoice.reference_id
#         ws[f"{get_column_letter(2)}{row_num}"] = invoice.customer.site_name if invoice.customer else ''
#         ws[f"{get_column_letter(3)}{row_num}"] = invoice.profile_name
#         ws[f"{get_column_letter(4)}{row_num}"] = invoice.order_number
#         ws[f"{get_column_letter(5)}{row_num}"] = invoice.repeat_every
#         ws[f"{get_column_letter(6)}{row_num}"] = invoice.start_date.strftime('%Y-%m-%d') if invoice.start_date else ''
#         ws[f"{get_column_letter(7)}{row_num}"] = invoice.end_date.strftime('%Y-%m-%d') if invoice.end_date else ''
#         ws[f"{get_column_letter(8)}{row_num}"] = invoice.sales_person.name if invoice.sales_person else ''
#         ws[f"{get_column_letter(9)}{row_num}"] = invoice.billing_address
#         ws[f"{get_column_letter(10)}{row_num}"] = invoice.gst_treatment

#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)

#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=recurring_invoices_export.xlsx'
#     response.write(output.read())

#     return response




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_recurring_invoice(request):
    serializer = RecurringInvoiceSerializer(data=request.data)
    if serializer.is_valid():
        recurring_invoice = serializer.save()
        return Response({
            "message": "Recurring Invoice added successfully!",
            "reference_id": recurring_invoice.reference_id,
            "billing_address": recurring_invoice.billing_address,
            "status": recurring_invoice.status
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_recurring_invoice(request, pk):
    try:
        recurring_invoice = RecurringInvoice.objects.get(pk=pk)
    except RecurringInvoice.DoesNotExist:
        return Response({"error": "Recurring Invoice not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = RecurringInvoiceSerializer(recurring_invoice, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Recurring Invoice updated successfully!",
            "reference_id": recurring_invoice.reference_id,
            "billing_address": recurring_invoice.billing_address,
            "status": recurring_invoice.status
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_recurring_invoice(request, pk):
    try:
        recurring_invoice = RecurringInvoice.objects.get(pk=pk)
        recurring_invoice.delete()
        return Response({"message": "Recurring Invoice deleted successfully!"}, status=status.HTTP_200_OK)
    except RecurringInvoice.DoesNotExist:
        return Response({"error": "Recurring Invoice not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def recurring_invoice_list(request):
    recurring_invoices = RecurringInvoice.objects.all()
    serializer = RecurringInvoiceSerializer(recurring_invoices, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_recurring_invoices_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Recurring Invoices"

    headers = [
        'Reference ID', 'Customer', 'Profile Name', 'Order Number', 'Repeat Every',
        'Start Date', 'End Date', 'Sales Person', 'Billing Address', 'GST Treatment', 'Status'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    recurring_invoices = RecurringInvoice.objects.all()

    for row_num, invoice in enumerate(recurring_invoices, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = invoice.reference_id
        ws[f"{get_column_letter(2)}{row_num}"] = invoice.customer.site_name if invoice.customer else ''
        ws[f"{get_column_letter(3)}{row_num}"] = invoice.profile_name
        ws[f"{get_column_letter(4)}{row_num}"] = invoice.order_number
        ws[f"{get_column_letter(5)}{row_num}"] = invoice.repeat_every
        ws[f"{get_column_letter(6)}{row_num}"] = invoice.start_date.strftime('%Y-%m-%d') if invoice.start_date else ''
        ws[f"{get_column_letter(7)}{row_num}"] = invoice.end_date.strftime('%Y-%m-%d') if invoice.end_date else ''
        ws[f"{get_column_letter(8)}{row_num}"] = invoice.sales_person.name if invoice.sales_person else ''
        ws[f"{get_column_letter(9)}{row_num}"] = invoice.billing_address
        ws[f"{get_column_letter(10)}{row_num}"] = invoice.gst_treatment
        ws[f"{get_column_letter(11)}{row_num}"] = invoice.status

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=recurring_invoices_export.xlsx'
    response.write(output.read())

    return response






@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_invoice_from_recurring(request, pk):
    from datetime import timedelta, date
    try:
        ri = RecurringInvoice.objects.get(pk=pk, status='active')
    except RecurringInvoice.DoesNotExist:
        return Response({"error": "Recurring invoice not found or inactive"}, status=status.HTTP_404_NOT_FOUND)

    today = date.today()
    if not ri.should_generate(today):
        return Response({"error": "Invoice not due yet"}, status=status.HTTP_400_BAD_REQUEST)

    # Create Invoice
    invoice = Invoice.objects.create(
        customer=ri.customer,
        start_date=ri.get_next_date(),
        due_date=ri.get_next_date() + timedelta(days=30),
        discount=0,
        payment_term='cash'
    )

    # Copy items
    for item in ri.items.all():
        InvoiceItem.objects.create(
            invoice=invoice,
            item=item.item,
            rate=item.rate,
            qty=item.qty,
            tax=item.tax
        )

    ri.last_generated_date = ri.get_next_date()
    ri.save()

    return Response({
        "message": f"Invoice {invoice.reference_id} generated successfully",
        "invoice_id": invoice.id,
        "reference_id": invoice.reference_id
    }, status=status.HTTP_201_CREATED)


###########################################Payment Received views#########################################

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_payment_received(request):
    serializer = PaymentReceivedSerializer(data=request.data)
    if serializer.is_valid():
        payment = serializer.save()
        return Response({
            "message": "Payment Received added successfully!",
            "payment_number": payment.payment_number
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_payment_received(request, pk):
    try:
        payment = PaymentReceived.objects.get(pk=pk)
    except PaymentReceived.DoesNotExist:
        return Response({"error": "Payment Received not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = PaymentReceivedSerializer(payment, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Payment Received updated successfully!",
            "payment_number": payment.payment_number
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_payment_received(request, pk):
    try:
        payment = PaymentReceived.objects.get(pk=pk)
        payment.delete()
        return Response({"message": "Payment Received deleted successfully!"}, status=status.HTTP_200_OK)
    except PaymentReceived.DoesNotExist:
        return Response({"error": "Payment Received not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def payment_received_list(request):
    payments = PaymentReceived.objects.all()
    serializer = PaymentReceivedSerializer(payments, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_payment_received_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments Received"

    headers = [
        'Payment Number', 'Customer', 'Invoice', 'Amount', 'Date',
        'Payment Type', 'Tax Deducted'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    payments = PaymentReceived.objects.all()

    for row_num, payment in enumerate(payments, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = payment.payment_number
        ws[f"{get_column_letter(2)}{row_num}"] = payment.customer.site_name if payment.customer else ''
        ws[f"{get_column_letter(3)}{row_num}"] = payment.invoice.reference_id if payment.invoice else ''
        ws[f"{get_column_letter(4)}{row_num}"] = str(payment.amount)
        ws[f"{get_column_letter(5)}{row_num}"] = payment.date.strftime('%Y-%m-%d') if payment.date else ''
        ws[f"{get_column_letter(6)}{row_num}"] = payment.payment_type
        ws[f"{get_column_letter(7)}{row_num}"] = payment.tax_deducted

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=payments_received_export.xlsx'
    response.write(output.read())

    return response