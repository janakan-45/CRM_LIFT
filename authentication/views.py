from rest_framework.decorators import api_view, permission_classes,parser_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from .models import Lift, FloorID, Brand, MachineType, MachineBrand, DoorType, DoorBrand, LiftType, ControllerBrand, Cabin, Type, Make, Unit, Item,Complaint, Employee,Profile
from .serializers import LiftSerializer, FloorIDSerializer, BrandSerializer, MachineTypeSerializer, MachineBrandSerializer, DoorTypeSerializer, DoorBrandSerializer, LiftTypeSerializer, ControllerBrandSerializer, CabinSerializer, TypeSerializer, MakeSerializer, UnitSerializer, ItemSerializer, UserRegistrationSerializer, UserLoginSerializer,ComplaintSerializer, EmployeeSerializer,ResetPasswordSerializer,ForgotPasswordSerializer,ProfileSerializer,ChangePasswordSerializer
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.core.mail import send_mail
from django.conf import settings
from datetime import datetime
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.contrib.auth import get_user_model
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
User = get_user_model()
from django.contrib.auth.models import Group
from rest_framework.permissions import BasePermission
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from rest_framework.parsers import MultiPartParser, FormParser


class IsOwner(BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.groups.filter(name='owner').exists()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_owner_status(request):
    user = request.user
    return Response({
        'message': f'User {user.username} is an owner',
        'is_owner': True
    }, status=status.HTTP_200_OK)


class IsEmployee(BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.groups.filter(name='employee').exists()
    

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsEmployee])
def check_employee_status(request):
    user = request.user
    return Response({
        'message': f'User {user.username} is an employee',
        'is_employee': True
    }, status=status.HTTP_200_OK)

# @api_view(['POST'])
# @permission_classes([AllowAny])
# def register(request):
#     serializer = UserRegistrationSerializer(data=request.data)
#     if serializer.is_valid():
#         user = serializer.save()
        
#         # Check if owner group exists, create if it doesn't
#         owner_group, created = Group.objects.get_or_create(name='owner')
        
#         # Assign user to owner group
#         user.groups.add(owner_group)
        
#         refresh = RefreshToken.for_user(user)

#         # Send confirmation email
#         subject = 'Welcome to ATOM LIFT - Registration Successful'
#         message = (
#             f"Dear {user.username},\n\n"
#             "Thank you for registering with ATOM LIFT!\n"
#             "Your account has been successfully created with owner privileges.\n\n"
#             "You can now log in using your email and password.\n"
#             "If you have any questions, feel free to contact us.\n\n"
#             "Best regards,\n"
#             "The ATOM LIFT Team"
#         )
#         from_email = f"{settings.EMAIL_SENDER_NAME} <{settings.EMAIL_HOST_USER}>"
#         recipient_list = [user.email]

#         try:
#             send_mail(
#                 subject=subject,
#                 message=message,
#                 from_email=from_email,
#                 recipient_list=recipient_list,
#                 fail_silently=False,
#             )
#         except Exception as e:
#             # Log the error if needed, but don't fail the registration
#             print(f"Failed to send email: {str(e)}")

#         return Response({
#             'refresh': str(refresh),
#             'access': str(refresh.access_token),
#             'user': serializer.data,
#             'message': 'Registration successful! A confirmation email has been sent to your email address.'
#         }, status=status.HTTP_201_CREATED)
    
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from .models import CustomUser
from .serializers import ForgotPasswordSerializer, ResetPasswordSerializer
from datetime import datetime

@api_view(['POST'])
@permission_classes([AllowAny])
def forgot_password(request):
    serializer = ForgotPasswordSerializer(data=request.data)
    if serializer.is_valid():
        email = serializer.validated_data['email']
        try:
            user = CustomUser.objects.get(email__iexact=email)
            # Check if the user is an OWNER
            if user.role == 'OWNER':
                return Response(
                    {'error': 'Password reset is not allowed for OWNER accounts.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except CustomUser.DoesNotExist:
            return Response(
                {'error': 'No user found with this email address.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        token_generator = PasswordResetTokenGenerator()
        token = token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        
        # Construct reset link (adjust the frontend URL as needed)
        reset_url = f"{settings.FRONTEND_URL}/reset-password/{uid}/{token}/"
        
        # Send password reset email
        subject = 'ATOM LIFT - Password Reset Request'
        message = (
            f"Dear {user.username},\n\n"
            "You have requested to reset your password for your ATOM LIFT account.\n"
            f"Please click the following link to reset your password:\n\n"
            f"{reset_url}\n\n"
            "This link is valid for 1 hour. If you did not request a password reset, please ignore this email.\n\n"
            "Best regards,\n"
            "The ATOM LIFT Team"
        )
        from_email = f"{settings.EMAIL_SENDER_NAME} <{settings.EMAIL_HOST_USER}>"
        recipient_list = [user.email]

        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=from_email,
                recipient_list=recipient_list,
                fail_silently=False,
            )
        except Exception as e:
            print(f"Failed to send password reset email: {str(e)}")
            return Response(
                {'error': 'Failed to send password reset email.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response(
            {'message': 'Password reset email sent successfully.'},
            status=status.HTTP_200_OK
        )
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([AllowAny])
def reset_password(request):
    serializer = ResetPasswordSerializer(data=request.data)
    if serializer.is_valid():
        uid = serializer.validated_data['uid']
        token = serializer.validated_data['token']
        password = serializer.validated_data['password']
        
        try:
            user_id = urlsafe_base64_decode(uid).decode()
            user = CustomUser.objects.get(pk=user_id)
        except (TypeError, ValueError, OverflowError, CustomUser.DoesNotExist):
            return Response(
                {'error': 'Invalid user ID.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        token_generator = PasswordResetTokenGenerator()
        if token_generator.check_token(user, token):
            user.set_password(password)
            user.save()

            # Send confirmation email
            subject = 'ATOM LIFT - Password Reset Successful'
            message = (
                f"Dear {user.username},\n\n"
                "Your password for your ATOM LIFT account has been successfully reset.\n"
                f"Reset Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S %Z')}\n"
                "If this was not you, please secure your account immediately or contact support.\n\n"
                "Best regards,\n"
                "The ATOM LIFT Team"
            )
            from_email = f"{settings.EMAIL_SENDER_NAME} <{settings.EMAIL_HOST_USER}>"
            recipient_list = [user.email]

            try:
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=from_email,
                    recipient_list=recipient_list,
                    fail_silently=False,
                )
            except Exception as e:
                print(f"Failed to send password reset confirmation email: {str(e)}")

            return Response(
                {'message': 'Password reset successfully. A confirmation email has been sent.'},
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                {'error': 'Invalid or expired token.'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
###############################################################################################
# from rest_framework.parsers import MultiPartParser, FormParser


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def profile(request):
#     try:
#         profile = request.user.profile
#     except Profile.DoesNotExist:
#         profile = Profile.objects.create(user=request.user)

#     serializer = ProfileSerializer(profile, context={'request': request})
#     return Response(serializer.data)


# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser, FormParser])
# def update_profile(request):
#     try:
#         profile = request.user.profile
#     except Profile.DoesNotExist:
#         profile = Profile.objects.create(user=request.user)

#     serializer = ProfileSerializer(
#         profile, data=request.data, partial=True, context={'request': request}
#     )
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data, status=status.HTTP_200_OK)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def change_password(request):
#     serializer = ChangePasswordSerializer(data=request.data, context={'request': request})
#     if serializer.is_valid():
#         user = request.user
#         user.set_password(serializer.validated_data['new_password'])
#         user.save()

#         # Send confirmation email
#         subject = 'ATOM LIFT - Password Changed Successfully'
#         message = (
#             f"Dear {user.username},\n\n"
#             "Your password for your ATOM LIFT account has been successfully changed.\n"
#             f"Change Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S %Z')}\n"
#             "If this was not you, please secure your account immediately or contact support.\n\n"
#             "Best regards,\n"
#             "The ATOM LIFT Team"
#         )
#         from_email = f"{settings.EMAIL_SENDER_NAME} <{settings.EMAIL_HOST_USER}>"
#         recipient_list = [user.email]

#         try:
#             send_mail(
#                 subject=subject,
#                 message=message,
#                 from_email=from_email,
#                 recipient_list=recipient_list,
#                 fail_silently=False,
#             )
#         except Exception as e:
#             print(f"Failed to send password change email: {str(e)}")

#         return Response({
#             'message': 'Password changed successfully. A confirmation email has been sent.'
#         }, status=status.HTTP_200_OK)
    
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from rest_framework.permissions import IsAdminUser
from rest_framework import status
from .serializers import CreateUserSerializer, PermissionUpdateSerializer
from .models import CustomUser
from django.core.mail import send_mail
from django.conf import settings
from datetime import datetime
from rest_framework.views import APIView
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from .models import CustomUser
from .serializers import UpdatePermissionsSerializer

class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']
            tokens = get_tokens_for_user(user)

            # Send login notification email
            login_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S %Z')
            subject = 'ATOM LIFT - Successful Login Notification'
            message = (
                f"Dear {user.username},\n\n"
                "You have successfully logged in to your ATOM LIFT account.\n\n"
                f"Login Time: {login_time}\n"
                "If this was not you, please secure your account immediately by changing your password.\n\n"
                "Best regards,\n"
                "The ATOM LIFT Team"
            )
            from_email = f"{settings.EMAIL_SENDER_NAME} <{settings.EMAIL_HOST_USER}>"
            recipient_list = [user.email]

            try:
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=from_email,
                    recipient_list=recipient_list,
                    fail_silently=False,
                )
            except Exception as e:
                # Log the error but don't fail the login
                print(f"Failed to send login email: {str(e)}")

            return Response({
                **tokens,
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "role": user.role,
                    "permissions": user.permissions  # include permissions
                },
                "message": "Login successful! A notification email has been sent to your email address."
            }, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_401_UNAUTHORIZED)


# Create Admin / Salesman (OWNER only)
class CreateUserView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        serializer = CreateUserSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response(CreateUserSerializer(user).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

# Edit User (OWNER only) - Supports partial updates
class EditUserView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        try:
            user = CustomUser.objects.get(pk=pk)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = CreateUserSerializer(user, data=request.data, partial=True)
        if serializer.is_valid():
            # Handle password update if provided
            password = serializer.validated_data.pop("password", None)
            if password:
                user.set_password(password)
            user = serializer.save()
            return Response(CreateUserSerializer(user).data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Delete User (OWNER only)
class DeleteUserView(APIView):
    permission_classes = [IsAdminUser]

    def delete(self, request, pk):
        try:
            user = CustomUser.objects.get(pk=pk)
            user.delete()
            return Response({"message": "User deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

# Assign Permissions (OWNER only)

class UpdatePermissionsView(APIView):
    permission_classes = [IsAdminUser]  # Only Owner/Admin can update

    def patch(self, request, pk):
        try:
            user = CustomUser.objects.get(pk=pk)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found"}, status=404)

        serializer = UpdatePermissionsSerializer(user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=200)
        return Response(serializer.errors, status=400)


class ListUsersView(APIView):
    permission_classes = [IsAdminUser]  # OWNER only

    def get(self, request):
        users = CustomUser.objects.exclude(role="OWNER")  # skip OWNER
        serializer = UserProfileSerializer(users, many=True)
        return Response(serializer.data)


class PermissionListView(APIView):
    permission_classes = [IsAdminUser]  # Only OWNER/Admin

    def get(self, request):
        permissions = [perm[0] for perm in CustomUser.PERMISSION_CHOICES]
        return Response({"permissions": permissions})


class ListPermissionsView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        return Response({"permissions": [p[0] for p in CustomUser.PERMISSION_CHOICES]})
    

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from .models import Profile
from .serializers import ProfileSerializer, ChangePasswordSerializer
from django.core.mail import send_mail
from django.conf import settings
from datetime import datetime

# Profile View (GET to retrieve profile)
class ProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            profile = request.user.profile
        except Profile.DoesNotExist:
            profile = Profile.objects.create(user=request.user)

        serializer = ProfileSerializer(profile, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

# Update Profile View (PUT to update profile with multipart support)
class UpdateProfileView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def put(self, request):
        try:
            profile = request.user.profile
        except Profile.DoesNotExist:
            profile = Profile.objects.create(user=request.user)

        serializer = ProfileSerializer(
            profile, data=request.data, partial=True, context={'request': request}
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Change Password View (POST to change password)
class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            user = request.user
            user.set_password(serializer.validated_data['new_password'])
            user.save()

            # Send confirmation email
            subject = 'ATOM LIFT - Password Changed Successfully'
            message = (
                f"Dear {user.username},\n\n"
                "Your password for your ATOM LIFT account has been successfully changed.\n"
                f"Change Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S %Z')}\n"
                "If this was not you, please secure your account immediately or contact support.\n\n"
                "Best regards,\n"
                "The ATOM LIFT Team"
            )
            from_email = f"{settings.EMAIL_SENDER_NAME} <{settings.EMAIL_HOST_USER}>"
            recipient_list = [user.email]

            try:
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=from_email,
                    recipient_list=recipient_list,
                    fail_silently=False,
                )
            except Exception as e:
                print(f"Failed to send password change email: {str(e)}")

            return Response({
                'message': 'Password changed successfully. A confirmation email has been sent.'
            }, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)





############################ Lift ######################################

@api_view(['POST'])
@permission_classes([IsAuthenticated, ])
def add_floor_id(request):
    serializer = FloorIDSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_floor_id(request, pk):
    try:
        floor_id = FloorID.objects.get(pk=pk)
    except FloorID.DoesNotExist:
        return Response({"error": "FloorID not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = FloorIDSerializer(floor_id, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "FloorID updated successfully!",
            "floor_id": floor_id.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_floor_id(request, pk):
    try:
        floor_id = FloorID.objects.get(pk=pk)
        floor_id.delete()
        return Response({"message": "FloorID deleted successfully!"}, status=status.HTTP_200_OK)
    except FloorID.DoesNotExist:
        return Response({"error": "FloorID not found"}, status=status.HTTP_404_NOT_FOUND)




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_brand(request):
    serializer = BrandSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_brand(request, pk):
    try:
        brand = Brand.objects.get(pk=pk)
    except Brand.DoesNotExist:
        return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = BrandSerializer(brand, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Brand updated successfully!",
            "brand_id": brand.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_brand(request, pk):
    try:
        brand = Brand.objects.get(pk=pk)
        brand.delete()
        return Response({"message": "Brand deleted successfully!"}, status=status.HTTP_200_OK)
    except Brand.DoesNotExist:
        return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_machine_type(request):
    serializer = MachineTypeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_machine_type(request, pk):
    try:
        machine_type = MachineType.objects.get(pk=pk)
    except MachineType.DoesNotExist:
        return Response({"error": "MachineType not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = MachineTypeSerializer(machine_type, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "MachineType updated successfully!",
            "machine_type_id": machine_type.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_machine_type(request, pk):
    try:
        machine_type = MachineType.objects.get(pk=pk)
        machine_type.delete()
        return Response({"message": "MachineType deleted successfully!"}, status=status.HTTP_200_OK)
    except MachineType.DoesNotExist:
        return Response({"error": "MachineType not found"}, status=status.HTTP_404_NOT_FOUND)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_machine_brand(request):
    serializer = MachineBrandSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_machine_brand(request, pk):
    try:
        machine_brand = MachineBrand.objects.get(pk=pk)
    except MachineBrand.DoesNotExist:
        return Response({"error": "MachineBrand not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = MachineBrandSerializer(machine_brand, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "MachineBrand updated successfully!",
            "machine_brand_id": machine_brand.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_machine_brand(request, pk):
    try:
        machine_brand = MachineBrand.objects.get(pk=pk)
        machine_brand.delete()
        return Response({"message": "MachineBrand deleted successfully!"}, status=status.HTTP_200_OK)
    except MachineBrand.DoesNotExist:
        return Response({"error": "MachineBrand not found"}, status=status.HTTP_404_NOT_FOUND)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_door_type(request):
    serializer = DoorTypeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_door_type(request, pk):
    try:
        door_type = DoorType.objects.get(pk=pk)
    except DoorType.DoesNotExist:
        return Response({"error": "DoorType not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = DoorTypeSerializer(door_type, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "DoorType updated successfully!",
            "door_type_id": door_type.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_door_type(request, pk):
    try:
        door_type = DoorType.objects.get(pk=pk)
        door_type.delete()
        return Response({"message": "DoorType deleted successfully!"}, status=status.HTTP_200_OK)
    except DoorType.DoesNotExist:
        return Response({"error": "DoorType not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_door_brand(request):
    serializer = DoorBrandSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_door_brand(request, pk):
    try:
        door_brand = DoorBrand.objects.get(pk=pk)
    except DoorBrand.DoesNotExist:
        return Response({"error": "DoorBrand not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = DoorBrandSerializer(door_brand, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "DoorBrand updated successfully!",
            "door_brand_id": door_brand.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_door_brand(request, pk):
    try:
        door_brand = DoorBrand.objects.get(pk=pk)
        door_brand.delete()
        return Response({"message": "DoorBrand deleted successfully!"}, status=status.HTTP_200_OK)
    except DoorBrand.DoesNotExist:
        return Response({"error": "DoorBrand not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_lift_type(request):
    serializer = LiftTypeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_lift_type(request, pk):
    try:
        lift_type = LiftType.objects.get(pk=pk)
    except LiftType.DoesNotExist:
        return Response({"error": "LiftType not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = LiftTypeSerializer(lift_type, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "LiftType updated successfully!",
            "lift_type_id": lift_type.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_lift_type(request, pk):
    try:
        lift_type = LiftType.objects.get(pk=pk)
        lift_type.delete()
        return Response({"message": "LiftType deleted successfully!"}, status=status.HTTP_200_OK)
    except LiftType.DoesNotExist:
        return Response({"error": "LiftType not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_controller_brand(request):
    serializer = ControllerBrandSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_controller_brand(request, pk):
    try:
        controller_brand = ControllerBrand.objects.get(pk=pk)
    except ControllerBrand.DoesNotExist:
        return Response({"error": "ControllerBrand not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = ControllerBrandSerializer(controller_brand, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "ControllerBrand updated successfully!",
            "controller_brand_id": controller_brand.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_controller_brand(request, pk):
    try:
        controller_brand = ControllerBrand.objects.get(pk=pk)
        controller_brand.delete()
        return Response({"message": "ControllerBrand deleted successfully!"}, status=status.HTTP_200_OK)
    except ControllerBrand.DoesNotExist:
        return Response({"error": "ControllerBrand not found"}, status=status.HTTP_404_NOT_FOUND)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_cabin(request):
    serializer = CabinSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_cabin(request, pk):
    try:
        cabin = Cabin.objects.get(pk=pk)
    except Cabin.DoesNotExist:
        return Response({"error": "Cabin not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = CabinSerializer(cabin, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Cabin updated successfully!",
            "cabin_id": cabin.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_cabin(request, pk):
    try:
        cabin = Cabin.objects.get(pk=pk)
        cabin.delete()
        return Response({"message": "Cabin deleted successfully!"}, status=status.HTTP_200_OK)
    except Cabin.DoesNotExist:
        return Response({"error": "Cabin not found"}, status=status.HTTP_404_NOT_FOUND)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_floor_ids(request):
    floor_ids = FloorID.objects.all()
    serializer = FloorIDSerializer(floor_ids, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_brands(request):
    brands = Brand.objects.all()
    serializer = BrandSerializer(brands, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_types(request):
    machine_types = MachineType.objects.all()
    serializer = MachineTypeSerializer(machine_types, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_brands(request):
    machine_brands = MachineBrand.objects.all()
    serializer = MachineBrandSerializer(machine_brands, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_door_types(request):
    door_types = DoorType.objects.all()
    serializer = DoorTypeSerializer(door_types, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_door_brands(request):
    door_brands = DoorBrand.objects.all()
    serializer = DoorBrandSerializer(door_brands, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_lift_types(request):
    lift_types = LiftType.objects.all()
    serializer = LiftTypeSerializer(lift_types, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_controller_brands(request):
    controller_brands = ControllerBrand.objects.all()
    serializer = ControllerBrandSerializer(controller_brands, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_cabins(request):
    cabins = Cabin.objects.all()
    serializer = CabinSerializer(cabins, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_lift(request):
    serializer = LiftSerializer(data=request.data)
    if serializer.is_valid():
        lift = serializer.save()
        return Response({
            "message": "Lift added successfully!",
            "lift_id": lift.id
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_lift(request, pk):
    try:
        lift = Lift.objects.get(pk=pk)
    except Lift.DoesNotExist:
        return Response({"error": "Lift not found"}, status=status.HTTP_404_NOT_FOUND)

    serializer = LiftSerializer(lift, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Lift updated successfully!",
            "lift_id": lift.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_lift(request, pk):
    try:
        lift = Lift.objects.get(pk=pk)
        lift.delete()
        return Response({"message": "Lift deleted successfully!"}, status=status.HTTP_200_OK)
    except Lift.DoesNotExist:
        return Response({"error": "Lift not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def lift_list(request):
    lifts = Lift.objects.all()
    serializer = LiftSerializer(lifts, many=True)
    return Response(serializer.data)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def export_lifts_to_excel(request):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Lifts"

#     headers = [
#         'Lift Code', 'Name', 'Price', 'Model', 'No of Passengers', 'Load (kg)', 'Speed',
#         'Floor ID', 'Brand', 'Lift Type', 'Machine Type', 'Machine Brand',
#         'Door Type', 'Door Brand', 'Controller Brand', 'Cabin'
#     ]

#     for col_num, header in enumerate(headers, 1):
#         ws[f"{get_column_letter(col_num)}1"] = header

#     lifts = Lift.objects.all()

#     for row_num, lift in enumerate(lifts, 2):
#         ws[f"{get_column_letter(1)}{row_num}"] = lift.lift_code
#         ws[f"{get_column_letter(2)}{row_num}"] = lift.name
#         ws[f"{get_column_letter(3)}{row_num}"] = float(lift.price)
#         ws[f"{get_column_letter(4)}{row_num}"] = lift.model
#         ws[f"{get_column_letter(5)}{row_num}"] = lift.no_of_passengers
#         ws[f"{get_column_letter(6)}{row_num}"] = lift.load_kg
#         ws[f"{get_column_letter(7)}{row_num}"] = lift.speed
#         ws[f"{get_column_letter(8)}{row_num}"] = lift.floor_id.value if lift.floor_id else ''
#         ws[f"{get_column_letter(9)}{row_num}"] = lift.brand.value if lift.brand else ''
#         ws[f"{get_column_letter(10)}{row_num}"] = lift.lift_type.value if lift.lift_type else ''
#         ws[f"{get_column_letter(11)}{row_num}"] = lift.machine_type.value if lift.machine_type else ''
#         ws[f"{get_column_letter(12)}{row_num}"] = lift.machine_brand.value if lift.machine_brand else ''
#         ws[f"{get_column_letter(13)}{row_num}"] = lift.door_type.value if lift.door_type else ''
#         ws[f"{get_column_letter(14)}{row_num}"] = lift.door_brand.value if lift.door_brand else ''
#         ws[f"{get_column_letter(15)}{row_num}"] = lift.controller_brand.value if lift.controller_brand else ''
#         ws[f"{get_column_letter(16)}{row_num}"] = lift.cabin.value if lift.cabin else ''

#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)

#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=lifts_export.xlsx'
#     response.write(output.read())

#     return response


# Updated views.py snippet for export_lifts_to_excel
# Add the new fields to the headers and data export in export_lifts_to_excel

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_lifts_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lifts"

    headers = [
        'Lift Code', 'Name', 'Price', 'Floor ID', 'Brand', 'Model',
        'No of Passengers', 'Load (kg)', 'Speed', 'Lift Type',
        'Machine Type', 'Machine Brand', 'Door Type', 'Door Brand',
        'Controller Brand', 'Cabin',
        # New fields added
        'Block', 'License No', 'License Start Date', 'License End Date'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    lifts = Lift.objects.all()

    for row_num, lift in enumerate(lifts, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = lift.lift_code
        ws[f"{get_column_letter(2)}{row_num}"] = lift.name
        ws[f"{get_column_letter(3)}{row_num}"] = str(lift.price)
        ws[f"{get_column_letter(4)}{row_num}"] = lift.floor_id.value if lift.floor_id else ''
        ws[f"{get_column_letter(5)}{row_num}"] = lift.brand.value if lift.brand else ''
        ws[f"{get_column_letter(6)}{row_num}"] = lift.model
        ws[f"{get_column_letter(7)}{row_num}"] = lift.no_of_passengers
        ws[f"{get_column_letter(8)}{row_num}"] = lift.load_kg
        ws[f"{get_column_letter(9)}{row_num}"] = lift.speed
        ws[f"{get_column_letter(10)}{row_num}"] = lift.lift_type.value if lift.lift_type else ''
        ws[f"{get_column_letter(11)}{row_num}"] = lift.machine_type.value if lift.machine_type else ''
        ws[f"{get_column_letter(12)}{row_num}"] = lift.machine_brand.value if lift.machine_brand else ''
        ws[f"{get_column_letter(13)}{row_num}"] = lift.door_type.value if lift.door_type else ''
        ws[f"{get_column_letter(14)}{row_num}"] = lift.door_brand.value if lift.door_brand else ''
        ws[f"{get_column_letter(15)}{row_num}"] = lift.controller_brand.value if lift.controller_brand else ''
        ws[f"{get_column_letter(16)}{row_num}"] = lift.cabin.value if lift.cabin else ''
        # New fields added
        ws[f"{get_column_letter(17)}{row_num}"] = lift.block or ''
        ws[f"{get_column_letter(18)}{row_num}"] = lift.license_no or ''
        ws[f"{get_column_letter(19)}{row_num}"] = lift.license_start_date.strftime('%Y-%m-%d') if lift.license_start_date else ''
        ws[f"{get_column_letter(20)}{row_num}"] = lift.license_end_date.strftime('%Y-%m-%d') if lift.license_end_date else ''

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=lifts_export.xlsx'
    response.write(output.read())

    return response
import csv
import io





import openpyxl

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser, FormParser])
# def import_lifts_csv(request):
#     """
#     Import Lift records from CSV or Excel (.xlsx)
#     Expected column order:
#     lift_code, name, price, floor_id_value, brand_value, model,
#     no_of_passengers, load_kg, speed, lift_type_value,
#     machine_type_value, machine_brand_value, door_type_value,
#     door_brand_value, controller_brand_value, cabin_value
#     """
#     if 'file' not in request.FILES:
#         return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

#     uploaded_file = request.FILES['file']
#     file_name = uploaded_file.name.lower()

#     if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
#         return Response({"error": "File must be CSV or Excel (.xlsx)"}, status=status.HTTP_400_BAD_REQUEST)

#     try:
#         rows = []
#         if file_name.endswith('.csv'):
#             decoded_file = uploaded_file.read().decode('utf-8-sig')
#             io_string = io.StringIO(decoded_file)
#             reader = csv.reader(io_string)
#             next(reader, None)
#             rows = list(reader)
#         else:
#             wb = openpyxl.load_workbook(uploaded_file)
#             ws = wb.active
#             rows = list(ws.iter_rows(min_row=2, values_only=True))

#         created_count = 0
#         for row in rows:
#             if not row or all(cell in [None, ''] for cell in row):
#                 continue

#             # Normalize passengers field
#             passengers_value = str(row[6]).strip() if row[6] else ''
#             if passengers_value and not passengers_value.lower().endswith("persons"):
#                 passengers_value = passengers_value.replace('.', '').strip()
#                 passengers_value = f"{passengers_value} Persons"

#             lift_data = {
#                 'lift_code': str(row[0]).strip() if row[0] else '',
#                 'name': str(row[1]).strip() if row[1] else '',
#                 'price': float(row[2]) if row[2] else 0.00,
#                 'model': str(row[5]).strip() if row[5] else '',
#                 'no_of_passengers': passengers_value,
#                 'load_kg': str(row[7]).strip() if row[7] else '',
#                 'speed': str(row[8]).strip() if row[8] else '',
#             }

#             # Foreign key mapping
#             fk_mapping = [
#                 ('floor_id', row[3], FloorID),
#                 ('brand', row[4], Brand),
#                 ('lift_type', row[9], LiftType),
#                 ('machine_type', row[10], MachineType),
#                 ('machine_brand', row[11], MachineBrand),
#                 ('door_type', row[12], DoorType),
#                 ('door_brand', row[13], DoorBrand),
#                 ('controller_brand', row[14], ControllerBrand),
#                 ('cabin', row[15], Cabin),
#             ]

#             for field, value, model_class in fk_mapping:
#                 if value and str(value).strip():
#                     instance, _ = model_class.objects.get_or_create(value=str(value).strip())
#                     lift_data[field] = instance.id
#                 else:
#                     lift_data[field] = None

#             serializer = LiftSerializer(data=lift_data)
#             if serializer.is_valid():
#                 serializer.save()
#                 created_count += 1
#             else:
#                 return Response({"error": serializer.errors, "row": row}, status=status.HTTP_400_BAD_REQUEST)

#         return Response(
#             {"message": f"{created_count} lifts imported successfully."},
#             status=status.HTTP_201_CREATED
#         )

#     except Exception as e:
#         return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_lifts_csv(request):
    """
    Import Lift records from CSV or Excel (.xlsx)
    Expected column order:
    lift_code, name, price, floor_id_value, brand_value, model,
    no_of_passengers, load_kg, speed, lift_type_value,
    machine_type_value, machine_brand_value, door_type_value,
    door_brand_value, controller_brand_value, cabin_value,
    block, license_no, license_start_date, license_end_date
    """
    if 'file' not in request.FILES:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

    uploaded_file = request.FILES['file']
    file_name = uploaded_file.name.lower()

    if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
        return Response({"error": "File must be CSV or Excel (.xlsx)"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = []
        if file_name.endswith('.csv'):
            decoded_file = uploaded_file.read().decode('utf-8-sig')
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader, None)
            rows = list(reader)
        else:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))

        created_count = 0
        for row in rows:
            if not row or all(cell in [None, ''] for cell in row):
                continue

            # --- Normalize passenger value ---
            passengers_value = str(row[6]).strip() if row[6] else ''
            if passengers_value:
                passengers_value = passengers_value.replace('.', '').strip()
                if not passengers_value.lower().endswith("persons"):
                    passengers_value = f"{passengers_value} Persons"

            # --- Date conversion helper ---
            def parse_date(value):
                if not value:
                    return None
                if isinstance(value, datetime):
                    return value.date()
                try:
                    return datetime.strptime(str(value), "%Y-%m-%d").date()
                except Exception:
                    try:
                        return datetime.strptime(str(value), "%d/%m/%Y").date()
                    except Exception:
                        return None

            # --- Build base lift data ---
            lift_data = {
                'lift_code': str(row[0]).strip() if row[0] else '',
                'name': str(row[1]).strip() if row[1] else '',
                'price': float(row[2]) if row[2] else 0.00,
                'model': str(row[5]).strip() if row[5] else '',
                'no_of_passengers': passengers_value,
                'load_kg': str(row[7]).strip() if row[7] else '',
                'speed': str(row[8]).strip() if row[8] else '',
                'block': str(row[16]).strip() if len(row) > 16 and row[16] else '',
                'license_no': str(row[17]).strip() if len(row) > 17 and row[17] else '',
                'license_start_date': parse_date(row[18]) if len(row) > 18 else None,
                'license_end_date': parse_date(row[19]) if len(row) > 19 else None,
            }

            # --- Handle foreign keys safely ---
            fk_mapping = [
                ('floor_id', row[3], FloorID),
                ('brand', row[4], Brand),
                ('lift_type', row[9], LiftType),
                ('machine_type', row[10], MachineType),
                ('machine_brand', row[11], MachineBrand),
                ('door_type', row[12], DoorType),
                ('door_brand', row[13], DoorBrand),
                ('controller_brand', row[14], ControllerBrand),
                ('cabin', row[15], Cabin),
            ]

            for field, value, model_class in fk_mapping:
                if value and str(value).strip():
                    instance, _ = model_class.objects.get_or_create(value=str(value).strip())
                    lift_data[field] = instance.id
                else:
                    lift_data[field] = None

            # --- Validate & Save ---
            serializer = LiftSerializer(data=lift_data)
            if serializer.is_valid():
                serializer.save()
                created_count += 1
            else:
                return Response({"error": serializer.errors, "row": row}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {"message": f"{created_count} lifts imported successfully."},
            status=status.HTTP_201_CREATED
        )

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




############################ Items ######################################

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_type(request):
    serializer = TypeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_type(request, pk):
    try:
        type_obj = Type.objects.get(pk=pk)
    except Type.DoesNotExist:
        return Response({"error": "Type not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = TypeSerializer(type_obj, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Type updated successfully!",
            "type_id": type_obj.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_type(request, pk):
    try:
        type_obj = Type.objects.get(pk=pk)
        type_obj.delete()
        return Response({"message": "Type deleted successfully!"}, status=status.HTTP_200_OK)
    except Type.DoesNotExist:
        return Response({"error": "Type not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_make(request):
    serializer = MakeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_make(request, pk):
    try:
        make = Make.objects.get(pk=pk)
    except Make.DoesNotExist:
        return Response({"error": "Make not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = MakeSerializer(make, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Make updated successfully!",
            "make_id": make.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_make(request, pk):
    try:
        make = Make.objects.get(pk=pk)
        make.delete()
        return Response({"message": "Make deleted successfully!"}, status=status.HTTP_200_OK)
    except Make.DoesNotExist:
        return Response({"error": "Make not found"}, status=status.HTTP_404_NOT_FOUND)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_unit(request):
    serializer = UnitSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_unit(request, pk):
    try:
        unit = Unit.objects.get(pk=pk)
    except Unit.DoesNotExist:
        return Response({"error": "Unit not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = UnitSerializer(unit, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Unit updated successfully!",
            "unit_id": unit.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_unit(request, pk):
    try:
        unit = Unit.objects.get(pk=pk)
        unit.delete()
        return Response({"message": "Unit deleted successfully!"}, status=status.HTTP_200_OK)
    except Unit.DoesNotExist:
        return Response({"error": "Unit not found"}, status=status.HTTP_404_NOT_FOUND)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_types(request):
    types = Type.objects.all()
    serializer = TypeSerializer(types, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_makes(request):
    makes = Make.objects.all()
    serializer = MakeSerializer(makes, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_units(request):
    units = Unit.objects.all()
    serializer = UnitSerializer(units, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_item(request):
    serializer = ItemSerializer(data=request.data)
    if serializer.is_valid():
        item = serializer.save()
        return Response({
            "message": "Item added successfully!",
            "item_id": item.id
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def item_list(request):
    items = Item.objects.all()
    serializer = ItemSerializer(items, many=True)
    return Response(serializer.data)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_item(request, pk):
    try:
        item = Item.objects.get(pk=pk)
    except Item.DoesNotExist:
        return Response({"error": "Item not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = ItemSerializer(item, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Item updated successfully!",
            "item_id": item.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_item(request, pk):
    try:
        item = Item.objects.get(pk=pk)
        item.delete()
        return Response({"message": "Item deleted successfully!"}, status=status.HTTP_200_OK)
    except Item.DoesNotExist:
        return Response({"error": "Item not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_items_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    headers = [
        'Item Number', 'Name', 'Make', 'Model', 'Type', 'Capacity', 'Threshold Qty',
        'Sale Price', 'Service Type', 'Tax Preference', 'Unit',
        'SAC Code', 'IGST', 'GST', 'Description'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    items = Item.objects.all()

    for row_num, item in enumerate(items, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = item.item_number
        ws[f"{get_column_letter(2)}{row_num}"] = item.name
        ws[f"{get_column_letter(3)}{row_num}"] = item.make.value if item.make else ''
        ws[f"{get_column_letter(4)}{row_num}"] = item.model
        ws[f"{get_column_letter(5)}{row_num}"] = item.type.value if item.type else ''
        ws[f"{get_column_letter(6)}{row_num}"] = item.capacity
        ws[f"{get_column_letter(7)}{row_num}"] = item.threshold_qty
        ws[f"{get_column_letter(8)}{row_num}"] = float(item.sale_price)
        ws[f"{get_column_letter(9)}{row_num}"] = item.service_type
        ws[f"{get_column_letter(10)}{row_num}"] = item.tax_preference
        ws[f"{get_column_letter(11)}{row_num}"] = item.unit.value if item.unit else ''
        ws[f"{get_column_letter(12)}{row_num}"] = item.sac_code if item.sac_code else ''
        ws[f"{get_column_letter(13)}{row_num}"] = float(item.igst) if item.igst else ''
        ws[f"{get_column_letter(14)}{row_num}"] = float(item.gst) if item.gst else ''
        ws[f"{get_column_letter(15)}{row_num}"] = item.description

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=items_export.xlsx'
    response.write(output.read())

    return response

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def import_items_csv(request):
#     if 'file' not in request.FILES:
#         return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

#     csv_file = request.FILES['file']
#     if not csv_file.name.endswith('.csv'):
#         return Response({"error": "File is not a CSV"}, status=status.HTTP_400_BAD_REQUEST)

#     try:
#         decoded_file = csv_file.read().decode('utf-8')
#         io_string = io.StringIO(decoded_file)
#         reader = csv.reader(io_string, delimiter=',')
#         next(reader)  # Skip header row
#         for row in reader:
#             if not row:  # Skip blank rows
#                 continue
#             if any(',' in field or not field.isalnum() for field in row):  # Check for commas or special characters
#                 return Response({"error": "CSV contains commas or special characters"}, status=status.HTTP_400_BAD_REQUEST)

#             # Map CSV columns to Item model fields
#             # Assuming CSV order: item_number, name, make_value, model, type_value, capacity,
#             # threshold_qty, sale_price, service_type, tax_preference, unit_value,
#             # sac_code, igst, gst, description
#             item_data = {
#                 'item_number': row[0],
#                 'name': row[1],
#                 'make': Make.objects.get_or_create(value=row[2])[0],
#                 'model': row[3],
#                 'type': Type.objects.get_or_create(value=row[4])[0],
#                 'capacity': row[5],
#                 'threshold_qty': int(row[6]) if row[6] else 0,
#                 'sale_price': float(row[7]) if row[7] else 0.00,
#                 'service_type': row[8] if row[8] in ['Services', 'Goods'] else 'Goods',
#                 'tax_preference': row[9] if row[9] in ['Non-Taxable', 'Taxable'] else 'Non-Taxable',
#                 'unit': Unit.objects.get_or_create(value=row[10])[0],
#                 'sac_code': row[11] if row[11] else None,
#                 'igst': float(row[12]) if row[12] else None,
#                 'gst': float(row[13]) if row[13] else None,
#                 'description': row[14] if row[14] else None,
#             }
#             serializer = ItemSerializer(data=item_data)
#             if serializer.is_valid():
#                 serializer.save()
#             else:
#                 return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

#         return Response({"message": "Items imported successfully!"}, status=status.HTTP_201_CREATED)

#     except Exception as e:
#         return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_items_csv(request):
    """
    Import Item records from CSV or Excel (.xlsx)
    Expected column order:
    item_number, name, make_value, model, type_value, capacity,
    threshold_qty, sale_price, service_type, tax_preference, unit_value,
    sac_code, igst, gst, description
    """
    if 'file' not in request.FILES:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

    uploaded_file = request.FILES['file']
    file_name = uploaded_file.name.lower()

    if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
        return Response({"error": "File must be CSV or Excel (.xlsx)"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # --- Read rows from file ---
        rows = []
        if file_name.endswith('.csv'):
            decoded_file = uploaded_file.read().decode('utf-8-sig')
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader, None)  # Skip header
            rows = list(reader)
        else:
            workbook = openpyxl.load_workbook(uploaded_file)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))

        created_count = 0
        skipped_count = 0

        for row in rows:
            # Skip completely empty rows
            if not row or all(cell in (None, '') for cell in row):
                skipped_count += 1
                continue

            try:
                # --- Safe helpers ---
                def to_str(v): return str(v).strip() if v else ''
                def to_int(v): 
                    try:
                        return int(float(v)) if v not in (None, '') else 0
                    except:
                        return 0
                def to_float(v):
                    try:
                        return float(v) if v not in (None, '') else 0.0
                    except:
                        return 0.0

                # --- Map columns to Item model fields ---
                item_data = {
                    'item_number': to_str(row[0]),
                    'name': to_str(row[1]),
                    'make': Make.objects.get_or_create(value=to_str(row[2]))[0].id if row[2] else None,
                    'model': to_str(row[3]),
                    'type': Type.objects.get_or_create(value=to_str(row[4]))[0].id if row[4] else None,
                    'capacity': to_str(row[5]),
                    'threshold_qty': to_int(row[6]),
                    'sale_price': to_float(row[7]),
                    'service_type': to_str(row[8]) if to_str(row[8]) in ['Services', 'Goods'] else 'Goods',
                    'tax_preference': to_str(row[9]) if to_str(row[9]) in ['Non-Taxable', 'Taxable'] else 'Non-Taxable',
                    'unit': Unit.objects.get_or_create(value=to_str(row[10]))[0].id if row[10] else None,
                    'sac_code': to_str(row[11]) if len(row) > 11 else '',
                    'igst': to_float(row[12]) if len(row) > 12 else 0.0,
                    'gst': to_float(row[13]) if len(row) > 13 else 0.0,
                    'description': to_str(row[14]) if len(row) > 14 else '',
                }

                serializer = ItemSerializer(data=item_data)
                if serializer.is_valid():
                    serializer.save()
                    created_count += 1
                else:
                    return Response(
                        {"error": serializer.errors, "row": row},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            except Exception as row_error:
                print(f"Row skipped due to error: {row_error}")
                skipped_count += 1
                continue

        return Response(
            {"message": f"Import complete. {created_count} items added, {skipped_count} skipped."},
            status=status.HTTP_201_CREATED
        )

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

####################################complaints########################################
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_employee(request):
    serializer = EmployeeSerializer(data=request.data)
    if serializer.is_valid():
        plain_password = request.data.get('password')
        employee = serializer.save()
        
        try:
            user = User.objects.create_user(
                username=employee.username,
                email=employee.email,
                password=plain_password
            )
            employee_group, created = Group.objects.get_or_create(name='employee')
            user.groups.add(employee_group)
        except Exception as e:
            employee.delete()
            return Response({"error": f"Failed to create user account: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        
        subject = 'Welcome to ATOM LIFT - Employee Account Created'
        message = (
            f"Dear {employee.name},\n\n"
            "Your employee account has been successfully created with ATOM LIFT.\n\n"
            "You can now log in using your email and password.\n"
            "If you have any questions, feel free to contact us.\n\n"
            "Best regards,\n"
            "The ATOM LIFT Team"
        )
        from_email = f"{settings.EMAIL_SENDER_NAME} <{settings.EMAIL_HOST_USER}>"
        recipient_list = [employee.email]

        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=from_email,
                recipient_list=recipient_list,
                fail_silently=False,
            )
        except Exception as e:
            print(f"Failed to send email: {str(e)}")

        return Response({
            "message": "Employee added successfully!",
            "employee_id": employee.id
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_employee(request, pk):
    try:
        employee = Employee.objects.get(pk=pk)
    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
    
    serializer = EmployeeSerializer(employee, data=request.data, partial=True)
    if serializer.is_valid():
        plain_password = request.data.get('password')
        employee = serializer.save()
        
        try:
            user = User.objects.get(email=employee.email)
            if 'username' in request.data:
                user.username = employee.username
            if 'email' in request.data:
                user.email = employee.email
            if plain_password:
                user.set_password(plain_password)
            user.save()
        except User.DoesNotExist:
            return Response({"error": "Associated user not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": f"Failed to update user account: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            "message": "Employee updated successfully!",
            "employee_id": employee.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_employee(request, pk):
    try:
        employee = Employee.objects.get(pk=pk)
        try:
            user = User.objects.get(email=employee.email)
            user.delete()
        except User.DoesNotExist:
            pass
        employee.delete()
        return Response({"message": "Employee deleted successfully!"}, status=status.HTTP_200_OK)
    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_employees(request):
    employees = Employee.objects.all()
    serializer = EmployeeSerializer(employees, many=True)
    return Response(serializer.data)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_complaint(request):
    serializer = ComplaintSerializer(data=request.data)
    if serializer.is_valid():
        complaint = serializer.save()
        return Response({
            "message": "Complaint added successfully!",
            "complaint_id": complaint.id
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def complaint_list(request):
    complaints = Complaint.objects.all()
    serializer = ComplaintSerializer(complaints, many=True)
    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_complaint(request, pk):
    try:
        complaint = Complaint.objects.get(pk=pk)
    except Complaint.DoesNotExist:
        return Response({"error": "Complaint not found"}, status=status.HTTP_404_NOT_FOUND)

    serializer = ComplaintSerializer(complaint, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Complaint updated successfully!",
            "complaint_id": complaint.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_complaint(request, pk):
    try:
        complaint = Complaint.objects.get(pk=pk)
        complaint.delete()
        return Response({"message": "Complaint deleted successfully!"}, status=status.HTTP_200_OK)
    except Complaint.DoesNotExist:
        return Response({"error": "Complaint not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_complaints_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints"

    headers = [
        'Reference', 'Type', 'Date', 'Customer Name', 'Contact Person Name',
        'Contact Person Mobile', 'Block/Wing', 'Assigned To', 'Priority',
        'Subject', 'Message'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    complaints = Complaint.objects.all()

    for row_num, complaint in enumerate(complaints, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = complaint.reference
        ws[f"{get_column_letter(2)}{row_num}"] = complaint.type
        ws[f"{get_column_letter(3)}{row_num}"] = complaint.date.strftime('%Y-%m-%d %H:%M:%S')
        ws[f"{get_column_letter(4)}{row_num}"] = complaint.customer.site_name if complaint.customer else ''
        ws[f"{get_column_letter(5)}{row_num}"] = complaint.contact_person_name
        ws[f"{get_column_letter(6)}{row_num}"] = complaint.contact_person_mobile
        ws[f"{get_column_letter(7)}{row_num}"] = complaint.block_wing
        ws[f"{get_column_letter(8)}{row_num}"] = complaint.assign_to.name if complaint.assign_to else ''
        ws[f"{get_column_letter(9)}{row_num}"] = complaint.priority
        ws[f"{get_column_letter(10)}{row_num}"] = complaint.subject
        ws[f"{get_column_letter(11)}{row_num}"] = complaint.message

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=complaints_export.xlsx'
    response.write(output.read())

    return response


from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import logging
from io import BytesIO

logger = logging.getLogger(__name__)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def print_complaint(request, pk):
    try:
        complaint = Complaint.objects.get(pk=pk)
    except Complaint.DoesNotExist:
        return Response({"error": "Complaint not found"}, status=status.HTTP_404_NOT_FOUND)

    try:
        # Prepare data
        context = {
            'company_name': 'Atom Lifts India Pvt Ltd',
            'address': 'No.87B, Pillayar Koll Street, Mannurpet, Ambattur Indus Estate, Chennai 50, CHENNAI',
            'phone': '9600087456',
            'email': 'admin@atomlifts.com',
            'ticket_no': complaint.reference,
            'ticket_date': complaint.date.strftime('%d/%m/%Y %H:%M:%S'),
            'ticket_type': complaint.type,
            'priority': complaint.priority,
            'customer_name': complaint.customer.site_name if complaint.customer else '',
            'site_address': complaint.customer.site_address if complaint.customer and hasattr(complaint.customer, 'site_address') else '',
            'contact_person': complaint.customer.contact_person_name,
            'contact_mobile': complaint.customer.phone,
            'block_wing': complaint.block_wing,
            'subject': complaint.subject,
            'message': complaint.message,
            'assigned_to': complaint.assign_to.name if complaint.assign_to else '',
            'customer_signature': complaint.customer_signature,
            'technician_remark': complaint.technician_remark,
            'technician_signature': complaint.technician_signature,
            'solution': complaint.solution
        }

        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
        styles = getSampleStyleSheet()
        story = []

        # Header
        header_style = ParagraphStyle(
            name='HeaderStyle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1  # Center
        )
        story.append(Paragraph(context['company_name'], header_style))
        story.append(Paragraph(context['address'], styles['Normal']))
        story.append(Paragraph(f"Phone: {context['phone']} | Email: {context['email']}", styles['Normal']))
        story.append(Spacer(1, 12))

        # Complaint Details Table
        data = [
            ['Ticket No:', context['ticket_no']],
            ['Date:', context['ticket_date']],
            ['Type:', context['ticket_type']],
            ['Priority:', context['priority']],
        ]
        table = Table(data, colWidths=[100, 400])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

        # Customer Details
        story.append(Paragraph("Customer Details", styles['Heading2']))
        data = [
            ['Customer Name:', context['customer_name']],
            ['Site Address:', context['site_address']],
            ['Contact Person:', context['contact_person']],
            ['Contact Mobile:', context['contact_mobile']],
            ['Block/Wing:', context['block_wing']],
        ]
        table = Table(data)
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

        # Complaint Details
        story.append(Paragraph("Complaint Details", styles['Heading2']))
        story.append(Paragraph(f"Subject: {context['subject']}", styles['Normal']))
        story.append(Paragraph(f"Message: {context['message']}", styles['Normal']))
        story.append(Paragraph(f"Assigned To: {context['assigned_to']}", styles['Normal']))
        story.append(Spacer(1, 12))

        # Resolution
        story.append(Paragraph("Resolution", styles['Heading2']))
        story.append(Paragraph(f"Technician Remark: {context['technician_remark']}", styles['Normal']))
        story.append(Paragraph(f"Solution: {context['solution']}", styles['Normal']))
        story.append(Spacer(1, 12))

        # Signatures (assuming base64 encoded images; adjust if different)
        if context['customer_signature']:
            story.append(Paragraph("Customer Signature", styles['Heading3']))
            # Note: ReportLab doesn't natively render base64 images easily; you'd need to decode first
            # For simplicity, we'll just show text placeholder
            story.append(Paragraph("Signature Placeholder", styles['Normal']))
        if context['technician_signature']:
            story.append(Paragraph("Technician Signature", styles['Heading3']))
            story.append(Paragraph("Signature Placeholder", styles['Normal']))

        # Build PDF
        doc.build(story)

        # Return PDF response
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=complaint_{context["ticket_no"]}.pdf'
        return response

    except Exception as e:
        logger.error(f"Unexpected error in print_complaint: {str(e)}")
        return Response({"error": f"Unexpected error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


    
# accounts/views.py
from rest_framework.permissions import AllowAny, IsAdminUser
from .models import CustomUser
from .serializers import RegisterSerializer, AdminApprovalSerializer, UserProfileSerializer, LoginSerializer
from .token import get_tokens_for_user

# ---------------- Register Admin Request ----------------
# class RegisterView(APIView):
#     permission_classes = [AllowAny]

#     def post(self, request):
#         serializer = RegisterSerializer(data=request.data)
#         if serializer.is_valid():
#             user = serializer.save()
#             tokens = get_tokens_for_user(user)
#             return Response({
#                 **tokens,
#                 "user": {
#                     "id": user.id,
#                     "username": user.username,
#                     "role": user.role
#                 }
#             }, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# ---------------- Login ---------------



# ---------------- Approve Admin (OWNER only) ----------------
# class ApproveAdminView(APIView):
#     permission_classes = [IsAdminUser]  # Only OWNER can approve

#     def patch(self, request, pk):
#         try:
#             user = CustomUser.objects.get(pk=pk)
#             if user.role == "PENDING":
#                 user.role = "ADMIN"
#                 user.save()
#                 return Response({
#                     "id": user.id,
#                     "username": user.username,
#                     "role": user.role
#                 }, status=status.HTTP_200_OK)
#             return Response({"error": "User is not pending approval"}, status=400)
#         except CustomUser.DoesNotExist:
#             return Response({"error": "User not found"}, status=404)

# ---------------- Create Salesman (OWNER only) ----------------
# class CreateSalesmanView(APIView):
#     permission_classes = [IsAdminUser]  # Only OWNER

#     def post(self, request):
#         username = request.data.get("username")
#         email = request.data.get("email")
#         password = request.data.get("password")

#         user = CustomUser.objects.create(
#             username=username,
#             email=email,
#             password=password,
#             role="SALESMAN"
#         )
#         tokens = get_tokens_for_user(user)
#         return Response({
#             **tokens,
#             "user": {
#                 "id": user.id,
#                 "username": user.username,
#                 "role": user.role
#             }
#         }, status=status.HTTP_201_CREATED)

# ---------------- List All Users ----------------
# class UserListView(APIView):
#     permission_classes = [IsAdminUser]

#     def get(self, request):
#         users = CustomUser.objects.all()
#         serializer = UserProfileSerializer(users, many=True)
#         return Response(serializer.data, status=status.HTTP_200_OK)
    




