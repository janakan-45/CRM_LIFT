from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from .models import AMC, AMCType, PaymentTerms,Item
from .serializers import AMCSerializer, AMCTypeSerializer, PaymentTermsSerializer,CustomerSerializer
import csv
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
from django.utils import timezone
from rest_framework.permissions import BasePermission
from rest_framework.parsers import MultiPartParser, FormParser
from datetime import datetime

###################################amc views######################################

class IsOwner(BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.groups.filter(name='owner').exists()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_owner_status(request):
    user = request.user
    return Response({
        'message': f'User {user.username} is an owner',
        'is_owner': True
    }, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_amc_type(request):
    serializer = AMCTypeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_amc_type(request, pk):
    try:
        amc_type = AMCType.objects.get(pk=pk)
    except AMCType.DoesNotExist:
        return Response({"error": "AMCType not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = AMCTypeSerializer(amc_type, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "AMCType updated successfully!",
            "amc_type_id": amc_type.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_amc_type(request, pk):
    try:
        amc_type = AMCType.objects.get(pk=pk)
        amc_type.delete()
        return Response({"message": "AMCType deleted successfully!"}, status=status.HTTP_200_OK)
    except AMCType.DoesNotExist:
        return Response({"error": "AMCType not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_payment_terms(request):
    serializer = PaymentTermsSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_payment_terms(request, pk):
    try:
        payment_terms = PaymentTerms.objects.get(pk=pk)
    except PaymentTerms.DoesNotExist:
        return Response({"error": "PaymentTerms not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = PaymentTermsSerializer(payment_terms, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "PaymentTerms updated successfully!",
            "payment_terms_id": payment_terms.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_payment_terms(request, pk):
    try:
        payment_terms = PaymentTerms.objects.get(pk=pk)
        payment_terms.delete()
        return Response({"message": "PaymentTerms deleted successfully!"}, status=status.HTTP_200_OK)
    except PaymentTerms.DoesNotExist:
        return Response({"error": "PaymentTerms not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_amc_types(request):
    amc_types = AMCType.objects.all()
    serializer = AMCTypeSerializer(amc_types, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_payment_terms(request):
    payment_terms = PaymentTerms.objects.all()
    serializer = PaymentTermsSerializer(payment_terms, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_amc(request):
    serializer = AMCSerializer(data=request.data)
    if serializer.is_valid():
        amc = serializer.save()
        return Response({
            "message": "AMC added successfully!",
            "reference_id": amc.reference_id,
            "customer_id": amc.customer.id,
            "customer_name": amc.customer.site_name
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_amc(request, pk):
    try:
        amc = AMC.objects.get(pk=pk)
    except AMC.DoesNotExist:
        return Response({"error": "AMC not found"}, status=status.HTTP_404_NOT_FOUND)

    serializer = AMCSerializer(amc, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "AMC updated successfully!",
            "reference_id": amc.reference_id,
            "customer_id": amc.customer.customer_id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_amc(request, pk):
    try:
        amc = AMC.objects.get(pk=pk)
        amc.delete()
        return Response({"message": "AMC deleted successfully!"}, status=status.HTTP_200_OK)
    except AMC.DoesNotExist:
        return Response({"error": "AMC not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def amc_list(request):
    amcs = AMC.objects.all()
    serializer = AMCSerializer(amcs, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_amc_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "AMCs"

    headers = [
        'Reference ID', 'Customer ID', 'Invoice Frequency', 'AMC Type', 'Payment Terms',
        'Start Date', 'End Date', 'Equipment No', 'Notes', 'Generate Contract',
        'No of Services', 'Price', 'No of Lifts', 'GST Percentage', 'Total',
        'Contract Amount', 'Total Amount Paid', 'Amount Due', 'Status',
        'AMC Service Item'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    amcs = AMC.objects.all()
    for row_num, amc in enumerate(amcs, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = amc.reference_id
        ws[f"{get_column_letter(2)}{row_num}"] = amc.customer.customer_id if amc.customer else ''
        ws[f"{get_column_letter(3)}{row_num}"] = amc.invoice_frequency
        ws[f"{get_column_letter(4)}{row_num}"] = amc.amc_type.name if amc.amc_type else ''
        ws[f"{get_column_letter(5)}{row_num}"] = amc.payment_terms.name if amc.payment_terms else ''
        ws[f"{get_column_letter(6)}{row_num}"] = amc.start_date
        ws[f"{get_column_letter(7)}{row_num}"] = amc.end_date
        ws[f"{get_column_letter(8)}{row_num}"] = amc.equipment_no
        ws[f"{get_column_letter(9)}{row_num}"] = amc.notes
        ws[f"{get_column_letter(10)}{row_num}"] = amc.is_generate_contract
        ws[f"{get_column_letter(11)}{row_num}"] = amc.no_of_services
        ws[f"{get_column_letter(12)}{row_num}"] = float(amc.price)
        ws[f"{get_column_letter(13)}{row_num}"] = amc.no_of_lifts
        ws[f"{get_column_letter(14)}{row_num}"] = float(amc.gst_percentage)
        ws[f"{get_column_letter(15)}{row_num}"] = float(amc.total)
        ws[f"{get_column_letter(16)}{row_num}"] = float(amc.contract_amount)
        ws[f"{get_column_letter(17)}{row_num}"] = float(amc.total_amount_paid)
        ws[f"{get_column_letter(18)}{row_num}"] = float(amc.amount_due)
        ws[f"{get_column_letter(19)}{row_num}"] = amc.status
        ws[f"{get_column_letter(20)}{row_num}"] = amc.amc_service_item.name if amc.amc_service_item else ''

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=amcs_export.xlsx'
    response.write(output.read())

    return response

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def import_amc_csv(request):
#     if 'file' not in request.FILES:
#         return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

#     csv_file = request.FILES['file']
#     if not csv_file.name.endswith('.csv'):
#         return Response({"error": "File is not a CSV"}, status=status.HTTP_400_BAD_REQUEST)

#     try:
#         decoded_file = csv_file.read().decode('utf-8')
#         io_string = io.StringIO(decoded_file)
#         reader = csv.reader(io_string, delimiter=',')
#         next(reader)  # Skip header row
#         for row in reader:
#             if not row:  # Skip blank rows
#                 continue
#             if any(',' in field for field in row):  # Check for commas in data
#                 return Response({"error": "CSV contains commas in data"}, status=status.HTTP_400_BAD_REQUEST)

#             # Map CSV columns to AMC model fields
#             amc_data = {
#                 'customer': CustomerSerializer().get_or_create_customer(customer_id=row[0])[0],  # Custom method to get or create customer
#                 'invoice_frequency': row[1] if row[1] in ['annually', 'semi_annually', 'quarterly', 'monthly', 'per_service'] else 'annually',
#                 'amc_type': AMCType.objects.get_or_create(name=row[2])[0],
#                 'payment_terms': PaymentTerms.objects.get_or_create(name=row[3])[0],
#                 'start_date': timezone.datetime.strptime(row[4], '%Y-%m-%d').date(),
#                 'end_date': timezone.datetime.strptime(row[5], '%Y-%m-%d').date(),
#                 'equipment_no': row[6] if row[6] else '',
#                 'notes': row[7] if row[7] else '',
#                 'is_generate_contract': row[8].lower() == 'true' if row[8] else False,
#                 'no_of_services': int(row[9]) if row[9] else 12,
#                 'price': float(row[10]) if row[10] else 0.00,
#                 'no_of_lifts': int(row[11]) if row[11] else 0,
#                 'gst_percentage': float(row[12]) if row[12] else 0.00,
#                 'amc_service_item': Item.objects.get_or_create(name=row[13])[0] if row[13] else None,
#             }
#             serializer = AMCSerializer(data=amc_data)
#             if serializer.is_valid():
#                 serializer.save()
#             else:
#                 return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

#         return Response({"message": "AMC details imported successfully!"}, status=status.HTTP_201_CREATED)

#     except Exception as e:
#         return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

import io
import openpyxl


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def import_amc_csv(request):
#     if 'file' not in request.FILES:
#         return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

#     uploaded_file = request.FILES['file']
#     file_name = uploaded_file.name.lower()
    
#     if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
#         return Response({"error": "File must be CSV or Excel (.xlsx)"}, status=status.HTTP_400_BAD_REQUEST)

#     try:
#         if file_name.endswith('.csv'):
#             decoded_file = uploaded_file.read().decode('utf-8')
#             io_string = io.StringIO(decoded_file)
#             reader = csv.reader(io_string, delimiter=',')
#             next(reader)  # Skip header row
#             rows = reader
#         else:  # Excel file
#             workbook = openpyxl.load_workbook(uploaded_file)
#             worksheet = workbook.active
#             rows = worksheet.iter_rows(min_row=2, values_only=True)  # Skip header row

#         for row in rows:
#             if not row or all(cell is None or cell == '' for cell in row):  # Skip blank rows
#                 continue
                
#             if file_name.endswith('.csv') and any(',' in str(field) for field in row if field):  # Check for commas in CSV data
#                 return Response({"error": "CSV contains commas in data"}, status=status.HTTP_400_BAD_REQUEST)

#             # Map columns to AMC model fields
#             amc_data = {
#                 'customer': CustomerSerializer().get_or_create_customer(row[0])[0],  # Custom method to get or create customer
#                 'invoice_frequency': row[1] if row[1] in ['annually', 'semi_annually', 'quarterly', 'monthly', 'per_service'] else 'annually',
#                 'amc_type': AMCType.objects.get_or_create(name=row[2])[0],
#                 'payment_terms': PaymentTerms.objects.get_or_create(name=row[3])[0],
#                 'start_date': timezone.datetime.strptime(str(row[4]), '%Y-%m-%d').date() if row[4] else None,
#                 'end_date': timezone.datetime.strptime(str(row[5]), '%Y-%m-%d').date() if row[5] else None,
#                 'equipment_no': str(row[6]) if row[6] else '',
#                 'notes': str(row[7]) if row[7] else '',
#                 'is_generate_contract': str(row[8]).lower() == 'true' if row[8] else False,
#                 'no_of_services': int(row[9]) if row[9] else 12,
#                 'price': float(row[10]) if row[10] else 0.00,
#                 'no_of_lifts': int(row[11]) if row[11] else 0,
#                 'gst_percentage': float(row[12]) if row[12] else 0.00,
#                 'amc_service_item': Item.objects.get_or_create(name=row[13])[0] if row[13] else None,
#             }
#             serializer = AMCSerializer(data=amc_data)
#             if serializer.is_valid():
#                 serializer.save()
#             else:
#                 return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

#         return Response({"message": "AMC details imported successfully!"}, status=status.HTTP_201_CREATED)

#     except Exception as e:
#         return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_amc_csv(request):
    """
    Import AMC data from CSV or Excel (.xlsx)
    Expected column order:
    customer_id, invoice_frequency, amc_type, payment_terms, start_date,
    end_date, equipment_no, notes, is_generate_contract, no_of_services,
    price, no_of_lifts, gst_percentage, amc_service_item
    """
    if 'file' not in request.FILES:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

    uploaded_file = request.FILES['file']
    file_name = uploaded_file.name.lower()

    if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
        return Response({"error": "File must be CSV or Excel (.xlsx)"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # --- Read rows ---
        if file_name.endswith('.csv'):
            decoded_file = uploaded_file.read().decode('utf-8-sig')
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader, None)  # Skip header row
            rows = list(reader)
        else:
            workbook = openpyxl.load_workbook(uploaded_file)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))  # Skip header row

        created_count = 0
        skipped_count = 0

        # --- Helper functions ---
        def to_str(v): return str(v).strip() if v else ''
        def to_int(v):
            try:
                return int(float(v))
            except:
                return 0
        def to_float(v):
            try:
                return float(v)
            except:
                return 0.0
        def parse_date(v):
            if not v:
                return None
            if isinstance(v, datetime):
                return v.date()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(str(v), fmt).date()
                except Exception:
                    continue
            return None

        for row in rows:
            if not row or all(cell in [None, ''] for cell in row):
                skipped_count += 1
                continue

            try:
                # --- Fetch or create related objects ---
                invoice_freq = to_str(row[1]).lower()
                if invoice_freq not in ['annually', 'semi_annually', 'quarterly', 'monthly', 'per_service']:
                    invoice_freq = 'annually'

                amc_type_obj = AMCType.objects.get_or_create(name=to_str(row[2]))[0] if row[2] else None
                payment_terms_obj = PaymentTerms.objects.get_or_create(name=to_str(row[3]))[0] if row[3] else None
                item_obj = Item.objects.get_or_create(name=to_str(row[13]))[0] if len(row) > 13 and row[13] else None

                # --- Try getting customer by ID or name ---
                customer_ref = to_str(row[0])
                customer = None
                if customer_ref.isdigit():
                    from sales.models import Customer
                    customer = Customer.objects.filter(id=int(customer_ref)).first()
                if not customer:
                    customer = Customer.objects.filter(site_name__iexact=customer_ref).first()
                if not customer:
                    skipped_count += 1
                    continue

                # --- Build AMC data ---
                amc_data = {
                    'customer': customer.id,
                    'invoice_frequency': invoice_freq,
                    'amc_type': amc_type_obj.id if amc_type_obj else None,
                    'payment_terms': payment_terms_obj.id if payment_terms_obj else None,
                    'start_date': parse_date(row[4]),
                    'end_date': parse_date(row[5]),
                    'equipment_no': to_str(row[6]),
                    'notes': to_str(row[7]),
                    'is_generate_contract': str(row[8]).lower() == 'true' if row[8] else False,
                    'no_of_services': to_int(row[9]) if len(row) > 9 else 12,
                    'price': to_float(row[10]),
                    'no_of_lifts': to_int(row[11]),
                    'gst_percentage': to_float(row[12]),
                    'amc_service_item': item_obj.id if item_obj else None,
                }

                serializer = AMCSerializer(data=amc_data)
                if serializer.is_valid():
                    serializer.save()
                    created_count += 1
                else:
                    return Response({"error": serializer.errors, "row": row}, status=status.HTTP_400_BAD_REQUEST)

            except Exception as row_error:
                print(f"Error importing row: {row_error}")
                skipped_count += 1
                continue

        return Response(
            {
                "message": f"AMC import completed. {created_count} added, {skipped_count} skipped.",
            },
            status=status.HTTP_201_CREATED,
        )

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
