from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import csv
from .models import Requisition
from .serializers import RequisitionSerializer


############################ Requisition ######################################

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_requisition(request):
    serializer = RequisitionSerializer(data=request.data)
    if serializer.is_valid():
        requisition = serializer.save()
        return Response({
            "message": "Requisition added successfully!",
            "requisition_id": requisition.id
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_requisition(request, pk):
    try:
        requisition = Requisition.objects.get(pk=pk)
    except Requisition.DoesNotExist:
        return Response({"error": "Requisition not found"}, status=status.HTTP_404_NOT_FOUND)
    serializer = RequisitionSerializer(requisition, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Requisition updated successfully!",
            "requisition_id": requisition.id
        }, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_requisition(request, pk):
    try:
        requisition = Requisition.objects.get(pk=pk)
        requisition.delete()
        return Response({"message": "Requisition deleted successfully!"}, status=status.HTTP_200_OK)
    except Requisition.DoesNotExist:
        return Response({"error": "Requisition not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def requisition_list(request):
    requisitions = Requisition.objects.all()
    serializer = RequisitionSerializer(requisitions, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_requisitions_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Requisitions"

    headers = [
        'Reference ID', 'Date', 'Item', 'Qty', 'Site', 'AMC ID', 'Service', 'Employee'
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    requisitions = Requisition.objects.all()

    for row_num, req in enumerate(requisitions, 2):
        ws[f"{get_column_letter(1)}{row_num}"] = req.reference_id
        ws[f"{get_column_letter(2)}{row_num}"] = req.date.strftime('%Y-%m-%d')
        ws[f"{get_column_letter(3)}{row_num}"] = req.item.value if req.item else ''
        ws[f"{get_column_letter(4)}{row_num}"] = req.qty
        ws[f"{get_column_letter(5)}{row_num}"] = req.site.site_name if req.site else ''
        ws[f"{get_column_letter(6)}{row_num}"] = req.amc_id.reference_id if req.amc_id else ''
        ws[f"{get_column_letter(7)}{row_num}"] = req.service
        ws[f"{get_column_letter(8)}{row_num}"] = req.employee.name if req.employee else ''

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=requisitions_export.xlsx'
    response.write(output.read())

    return response

